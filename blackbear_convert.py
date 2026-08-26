#!/usr/bin/env python3
"""
Black Bear Burger — Deliveroo "Items Sold" → Supy sales-upload converter.

The client emails a single multi-tab .xlsx exported from the Deliveroo Looker
report ("Items Sold <day> <Mon>"). Each tab is one sales date; columns are
grouped by Deliveroo restaurant, rows are menu items.

Several Deliveroo restaurants map onto a single Supy branch (the 20Ft Fried
Chicken virtual brand shares a kitchen with the Black Bear Burger site), so the
converter merges them by summing per item name — see
mappings/blackbear_branches.csv.

Output: one Supy-format .xlsx per Supy branch, covering every date found,
written to output/blackbear/<run-date>/.

Usage:
    python blackbear_convert.py "Items Sold 17th Aug.xlsx"
    python blackbear_convert.py in.xlsx --out-dir output/bbb --vat-rate 0.20
    python blackbear_convert.py in.xlsx --one-file-per-date
"""

from __future__ import annotations

import argparse
import csv
import os
import pathlib
import re
import smtplib
import sys
from collections import defaultdict
from datetime import date, datetime
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

import openpyxl
import pandas as pd
from dotenv import load_dotenv

load_dotenv()

# ── Supy upload contract ────────────────────────────────────────────────
SUPY_COLUMNS = [
    "Sales Date *",
    "POS Item ID *",
    "POS Item Name",
    "Sold QTY *",
    "Total Discount Value",
    "Total sales excl. tax *",
    "Total sales incl. tax *",
    "Order ID",
    "Sales Type Code",
    "Parent Item ID",
]
SUPY_DATE_FORMAT = "%d-%b-%Y"

# ── Deliveroo export contract ───────────────────────────────────────────
RESTAURANT_HEADER = "Restaurant Information Restaurant Name"
ITEM_HEADER = "Menu Item Name"
QTY_METRIC = "Count Orders (incl Undelivered)"
GROSS_METRIC = "Item Value Sum (before discounts)"
NET_METRIC = "Item Value Sum (after item discounts)"

SHEET_DATE_RE = re.compile(r"(\d{1,2})\s*(?:st|nd|rd|th)?\s+([A-Za-z]{3})", re.I)
MONTHS = {m.lower(): i for i, m in enumerate(
    ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
     "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"], start=1)}

DEFAULT_MAPPING = pathlib.Path(__file__).parent / "mappings" / "blackbear_branches.csv"


class ConvertError(Exception):
    pass


# ── Mapping ─────────────────────────────────────────────────────────────
def load_mapping(path: pathlib.Path) -> dict[str, str]:
    if not path.exists():
        raise ConvertError(f"Branch mapping not found: {path}")
    mapping: dict[str, str] = {}
    with path.open(newline="", encoding="utf-8") as fh:
        for row in csv.DictReader(fh):
            src = (row.get("deliveroo_restaurant") or "").strip()
            dst = (row.get("supy_branch") or "").strip()
            if src and dst:
                mapping[src.casefold()] = dst
    if not mapping:
        raise ConvertError(f"Branch mapping is empty: {path}")
    return mapping


# ── Sheet parsing ───────────────────────────────────────────────────────
def parse_sheet_date(title: str, year_hint: int | None, today: date) -> date | None:
    m = SHEET_DATE_RE.search(title)
    if not m:
        return None
    day = int(m.group(1))
    month = MONTHS.get(m.group(2).lower())
    if not month:
        return None
    if year_hint:
        year = year_hint
    else:
        # Sheets are recent history: a month ahead of today belongs to last year.
        year = today.year if month <= today.month else today.year - 1
    try:
        return date(year, month, day)
    except ValueError:
        return None


def _num(value) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(",", "").strip())
    except ValueError:
        return None


def read_sheet(ws) -> tuple[list[dict], str, list[dict]]:
    """Return (records, layout, footer) for one Deliveroo tab.

    Layout "full"  — 3 metric columns per restaurant (qty + gross + net).
    Layout "qty"   — 1 column per restaurant, counts only, no revenue.

    Looker appends a grand-total row at the bottom; it is stripped from the
    records and returned separately as `footer` for reconciliation.
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], "empty", []

    header = rows[0]
    second = rows[1] if len(rows) > 1 else ()
    has_metric_row = bool(second) and str(second[0] or "").strip() == ITEM_HEADER

    # Forward-fill the merged restaurant header.
    names: list[str | None] = []
    current = None
    for cell in header[1:]:
        text = str(cell).strip() if cell is not None else ""
        if text:
            current = text
        names.append(current)

    if has_metric_row:
        layout = "full"
        metrics = [str(c).strip() if c is not None else "" for c in second[1:]]
        data_rows = rows[2:]
    else:
        layout = "qty"
        metrics = [QTY_METRIC] * len(names)
        data_rows = rows[1:]

    # column index (0-based within the metric slice) → (restaurant, metric)
    slots: dict[int, tuple[str, str]] = {}
    for i, (name, metric) in enumerate(zip(names, metrics)):
        if name and metric:
            slots[i] = (name, metric)

    # Strip Looker's grand-total footer (always the final row) and keep it for
    # reconciliation. Only the last row qualifies, so a menu item literally
    # named "Total" higher up is left alone.
    footer_row = None
    while data_rows and (not data_rows[-1] or data_rows[-1][0] is None
                         or str(data_rows[-1][0]).strip() == ""):
        data_rows = data_rows[:-1]
    if data_rows and str(data_rows[-1][0]).strip().casefold() == "total":
        footer_row = data_rows[-1]
        data_rows = data_rows[:-1]

    def accumulate(rows_in):
        buckets: dict[tuple[str, str], dict[str, float | None]] = {}
        for row in rows_in:
            item = str(row[0]).strip() if row and row[0] is not None else ""
            if not item:
                continue
            values = row[1:]
            for i, (restaurant, metric) in slots.items():
                if i >= len(values):
                    continue
                val = _num(values[i])
                if val is None:
                    continue
                key = (restaurant, item)
                bucket = buckets.setdefault(key, {"qty": None, "gross": None, "net": None})
                field = {QTY_METRIC: "qty", GROSS_METRIC: "gross",
                         NET_METRIC: "net"}.get(metric)
                if field is None:
                    continue
                bucket[field] = (bucket[field] or 0.0) + val
        return [{"restaurant": r, "item": i, **vals} for (r, i), vals in buckets.items()]

    records = accumulate(data_rows)
    footer = accumulate([footer_row]) if footer_row else []
    return records, layout, footer


# ── Transform ───────────────────────────────────────────────────────────
def build_rows(records: list[dict], sales_date: date, vat_rate: float,
               mapping: dict[str, str], merge_notes: list[dict],
               unmapped: set[str]) -> dict[str, list[dict]]:
    """Merge Deliveroo restaurants into Supy branches → Supy rows per branch."""
    merged: dict[tuple[str, str], dict] = {}
    collisions: dict[tuple[str, str], set[str]] = defaultdict(set)

    for rec in records:
        branch = mapping.get(rec["restaurant"].casefold())
        if not branch:
            unmapped.add(rec["restaurant"])
            continue
        key = (branch, rec["item"])
        collisions[key].add(rec["restaurant"])
        acc = merged.setdefault(key, {"qty": 0.0, "gross": 0.0, "net": 0.0})
        for field in ("qty", "gross", "net"):
            if rec[field] is not None:
                acc[field] += rec[field]

    by_branch: dict[str, list[dict]] = defaultdict(list)
    date_str = sales_date.strftime(SUPY_DATE_FORMAT)
    for (branch, item), acc in merged.items():
        # Deliveroo's "Item Value Sum" figures are NET of VAT, so the after-discount
        # value IS the excl.-tax figure and VAT is added on top to get incl. tax.
        net = round(acc["net"], 2)
        gross = round(acc["gross"], 2)
        discount = round(max(gross - net, 0.0), 2)
        excl = net
        incl = round(net * (1.0 + vat_rate), 2)
        by_branch[branch].append({
            "Sales Date *": date_str,
            "POS Item ID *": item,
            "POS Item Name": item,
            "Sold QTY *": int(round(acc["qty"])),
            "Total Discount Value": discount,
            "Total sales excl. tax *": excl,
            "Total sales incl. tax *": incl,
            "Order ID": "",
            "Sales Type Code": "",
            "Parent Item ID": "",
        })

    per_branch_merges: dict[str, set[str]] = defaultdict(set)
    merged_items: dict[str, int] = defaultdict(int)
    for (branch, _item), sources in collisions.items():
        if len(sources) > 1:
            per_branch_merges[branch] |= sources
            merged_items[branch] += 1
    for branch, sources in per_branch_merges.items():
        merge_notes.append({
            "date": date_str,
            "branch": branch,
            "items": merged_items[branch],
            "sources": sorted(sources),
        })
    return by_branch


def reconcile(sheet_title: str, records: list[dict], footer: list[dict],
              tolerance: float = 0.05) -> list[str]:
    """Compare per-restaurant sums against Looker's own grand-total row."""
    if not footer:
        return [f"{sheet_title}: no grand-total row found — totals unverified"]
    sums: dict[str, dict[str, float]] = defaultdict(lambda: {"qty": 0.0, "gross": 0.0, "net": 0.0})
    for rec in records:
        for field in ("qty", "gross", "net"):
            if rec[field] is not None:
                sums[rec["restaurant"]][field] += rec[field]

    issues = []
    for exp in footer:
        got = sums.get(exp["restaurant"])
        if got is None:
            issues.append(f"{sheet_title}: '{exp['restaurant']}' has a total row but no item rows")
            continue
        # Only money is checked. Looker's grand-total for
        # "Count Orders (incl Undelivered)" is a DISTINCT order count, so it is
        # deliberately smaller than the sum of the per-item counts.
        for field, label in (("gross", "gross value"), ("net", "net value")):
            want = exp[field]
            if want is None:
                continue
            if abs(got[field] - want) > tolerance:
                issues.append(
                    f"{sheet_title}: '{exp['restaurant']}' {label} mismatch — "
                    f"items sum to {got[field]:,.2f}, Deliveroo total row says {want:,.2f}"
                )
    return issues


def safe_name(text: str) -> str:
    return re.sub(r"[^A-Za-z0-9]+", "_", text).strip("_")


def sheet_name(branch: str, taken: set[str]) -> str:
    """Excel tab name: <=31 chars, no []:*?/\\, unique within the workbook."""
    base = re.sub(r"[\[\]:*?/\\]", "-", branch).strip()[:31].strip() or "Branch"
    name, n = base, 2
    while name.casefold() in taken:
        suffix = f"~{n}"
        name = base[:31 - len(suffix)] + suffix
        n += 1
    taken.add(name.casefold())
    return name


def write_combined(path: pathlib.Path, collected: dict, summary: list[dict],
                   report: list[str]) -> None:
    """One workbook: Read Me + Summary + a tab per Supy branch."""
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        pd.DataFrame({"Black Bear Burger — Supy sales upload": report}).to_excel(
            writer, sheet_name="Read Me", index=False)
        pd.DataFrame(summary).to_excel(writer, sheet_name="Summary", index=False)

        taken = {"read me", "summary"}
        for branch in sorted(collected):
            per_date = collected[branch]
            rows = [r for d in sorted(per_date) for r in per_date[d]]
            df = pd.DataFrame(rows, columns=SUPY_COLUMNS)
            df = sort_supy(df)
            df.to_excel(writer, sheet_name=sheet_name(branch, taken), index=False)

        for ws in writer.book.worksheets:
            widths = {"Read Me": [96], "Summary": [34, 14, 9, 11, 11, 16, 16]}.get(
                ws.title, [14, 34, 34, 11, 20, 22, 22, 10, 16, 15])
            for i, width in enumerate(widths, start=1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
            ws.freeze_panes = "A2"


def sort_supy(df: pd.DataFrame) -> pd.DataFrame:
    return df.sort_values(
        ["Sales Date *", "POS Item Name"],
        key=lambda s: pd.to_datetime(s, format=SUPY_DATE_FORMAT, errors="coerce")
        if s.name == "Sales Date *" else s,
    ).reset_index(drop=True)


# ── Main ────────────────────────────────────────────────────────────────
def send_email(attachment: pathlib.Path, recipients: list[str], row_count: int,
               span: str, branches: int, report: pathlib.Path) -> None:
    """Send the combined POS workbook via Gmail SMTP."""
    gmail_user = os.environ.get("GMAIL_USER", "")
    gmail_password = os.environ.get("GMAIL_APP_PASSWORD", "")
    if not gmail_user or not gmail_password:
        raise ConvertError(
            "GMAIL_USER and GMAIL_APP_PASSWORD must be set in .env to send email.")

    body = (
        "Hi,\n\n"
        "Attached is the Black Bear Burger Deliveroo sales data converted into the "
        "Supy upload format.\n\n"
        f"  * Period      : {span}\n"
        f"  * Branches    : {branches} (one tab per Supy branch)\n"
        f"  * Rows        : {row_count:,}\n"
        f"  * Workbook    : {attachment.name}\n\n"
        "The 20Ft Fried Chicken restaurants have been merged into their paired Black "
        "Bear Burger branch, summed by item name per date. Every per-restaurant gross "
        "and net total reconciles exactly against Deliveroo's own grand-total row.\n\n"
        "Please read the 'Read Me' tab before uploading — it lists the assumptions "
        "that still need confirming (notably that Deliveroo's only quantity metric "
        "counts orders containing an item, not units sold).\n\n"
        "Regards,\nSupy POS Integration"
    )

    msg = MIMEMultipart()
    msg["From"] = gmail_user
    msg["To"] = ", ".join(recipients)
    msg["Subject"] = f"Black Bear Burger — Supy POS upload ({span})"
    msg.attach(MIMEText(body, "plain"))

    for path, subtype in ((attachment, "vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
                          (report, "plain")):
        with open(path, "rb") as fh:
            part = MIMEApplication(fh.read(), _subtype=subtype)
        part.add_header("Content-Disposition", "attachment", filename=path.name)
        msg.attach(part)

    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465, timeout=30) as server:
            server.login(gmail_user, gmail_password)
            server.sendmail(gmail_user, recipients, msg.as_string())
    except Exception as exc:
        raise ConvertError(f"Failed to send email: {exc}") from exc

    print(f"  ✓ Email sent → {', '.join(recipients)}")


def convert(src: pathlib.Path, out_dir: pathlib.Path, mapping_path: pathlib.Path,
            vat_rate: float, year_hint: int | None, one_file_per_date: bool,
            today: date):
    mapping = load_mapping(mapping_path)
    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)

    merge_notes: list[dict] = []
    recon: list[str] = []
    unmapped: set[str] = set()
    skipped_sheets: list[str] = []
    # branch -> date -> rows
    collected: dict[str, dict[date, list[dict]]] = defaultdict(dict)

    for ws in wb.worksheets:
        sales_date = parse_sheet_date(ws.title, year_hint, today)
        if sales_date is None:
            skipped_sheets.append(f"{ws.title} (no date in tab name)")
            continue

        records, layout, footer = read_sheet(ws)
        if layout == "empty" or not records:
            skipped_sheets.append(f"{ws.title} (no data rows)")
            continue
        if layout == "qty":
            skipped_sheets.append(
                f"{ws.title} (quantities only — no 'Item Value Sum' columns, "
                f"cannot produce required sales values)"
            )
            continue

        recon.extend(reconcile(ws.title, records, footer))

        for branch, rows in build_rows(records, sales_date, vat_rate,
                                       mapping, merge_notes, unmapped).items():
            collected[branch][sales_date] = rows

    if not collected:
        raise ConvertError("No convertible sheets found in the workbook.")

    out_dir.mkdir(parents=True, exist_ok=True)
    summary: list[dict] = []
    written = 0

    for branch in sorted(collected):
        per_date = collected[branch]
        dates = sorted(per_date)

        if one_file_per_date:
            groups = [(d, per_date[d]) for d in dates]
        else:
            all_rows = [r for d in dates for r in per_date[d]]
            groups = [(None, all_rows)]

        for tag, rows in groups:
            df = sort_supy(pd.DataFrame(rows, columns=SUPY_COLUMNS))

            if tag is None:
                span = (f"{dates[0]:%Y-%m-%d}_to_{dates[-1]:%Y-%m-%d}"
                        if len(dates) > 1 else f"{dates[0]:%Y-%m-%d}")
            else:
                span = f"{tag:%Y-%m-%d}"
            path = out_dir / f"{safe_name(branch)}_{span}.xlsx"
            df.to_excel(path, index=False)
            written += 1
            print(f"  ✓ {path.name:58s} {len(df):5d} rows  "
                  f"incl.tax £{df['Total sales incl. tax *'].sum():,.2f}")

        for d in dates:
            rows = per_date[d]
            summary.append({
                "Supy Branch": branch,
                "Sales Date": d.strftime(SUPY_DATE_FORMAT),
                "Items": len(rows),
                "Sold QTY": sum(r["Sold QTY *"] for r in rows),
                "Discount": round(sum(r["Total Discount Value"] for r in rows), 2),
                "Sales excl. tax": round(sum(r["Total sales excl. tax *"] for r in rows), 2),
                "Sales incl. tax": round(sum(r["Total sales incl. tax *"] for r in rows), 2),
            })

    summary_path = out_dir / "_summary.csv"
    pd.DataFrame(summary).to_excel(out_dir / "_summary.xlsx", index=False)
    pd.DataFrame(summary).to_csv(summary_path, index=False)

    report = [
        f"Source          : {src}",
        f"Branch mapping  : {mapping_path}",
        f"VAT rate        : {vat_rate:.2%} (Total sales incl. tax = excl. tax x {1 + vat_rate:.2f})",
        f"Branch files    : {written}",
        f"Rows total      : {sum(s['Items'] for s in summary)}",
        "",
        "ASSUMPTIONS — confirm before uploading:",
        "  * 'POS Item ID *' is filled with the Deliveroo menu item name. The Deliveroo",
        "    Looker export carries no item ID; if Supy expects real POS IDs, supply an",
        "    item master and remap this column.",
        "  * 'Sold QTY *' comes from 'Count Orders (incl Undelivered)'. That metric",
        "    counts ORDERS CONTAINING the item, not units sold — two burgers on one",
        "    order count as 1 — and it includes undelivered orders. Ask the client to",
        "    add a quantity measure to the Looker look for a true units figure.",
        "  * 'Total sales excl. tax *' = 'Item Value Sum (after item discounts)';",
        "    'Total Discount Value' = before-discount minus after-discount.",
        f"  * 'Total sales incl. tax *' is derived by multiplying by {1 + vat_rate:.2f};",
        "    Deliveroo does not export a VAT-inclusive figure.",
        "  * Modifier/option rows (e.g. 'Standard Cooking', 'Add Bacon', sauces) are",
        "    kept. Their values are supplements, not duplicates of the parent item —",
        "    branch revenue reconciles exactly only when they are included. Deliveroo",
        "    exports no parent/child link, so 'Parent Item ID' is left empty.",
    ]
    if unmapped:
        report += ["", "UNMAPPED DELIVEROO RESTAURANTS (excluded from output):"]
        report += [f"  - {n}" for n in sorted(unmapped)]
    if skipped_sheets:
        report += ["", "SKIPPED TABS:"]
        report += [f"  - {s}" for s in skipped_sheets]

    if merge_notes:
        rolled: dict[str, tuple[set[str], int, int]] = {}
        for note in merge_notes:
            srcs, items, days = rolled.get(note["branch"], (set(), 0, 0))
            rolled[note["branch"]] = (srcs | set(note["sources"]),
                                      items + note["items"], days + 1)
        report += ["", "MERGED BRANCHES (rows summed by item name):"]
        for branch in sorted(rolled):
            srcs, items, days = rolled[branch]
            report.append(f"  - {branch} ← {' + '.join(sorted(srcs))}")
            report.append(f"      {items} item rows combined across {days} dates")

    report += ["", "RECONCILIATION vs Deliveroo grand-total rows (money only):"]
    report += ([f"  ! {i}" for i in recon] if recon
               else ["  every per-restaurant gross and net total matches exactly"])

    (out_dir / "_report.txt").write_text("\n".join(report) + "\n", encoding="utf-8")

    all_dates = sorted({d for per_date in collected.values() for d in per_date})
    span = (f"{all_dates[0]:%Y-%m-%d}_to_{all_dates[-1]:%Y-%m-%d}"
            if len(all_dates) > 1 else f"{all_dates[0]:%Y-%m-%d}")
    combined_path = out_dir / f"BlackBearBurger_Supy_POS_Upload_{span}.xlsx"
    write_combined(combined_path, collected, summary, report)
    print(f"\n  ✓ {combined_path.name}  (Read Me + Summary + {len(collected)} branch tabs)")

    print("\n" + "\n".join(report))
    print(f"\nOutput directory: {out_dir}")
    return combined_path, sum(s["Items"] for s in summary), span.replace("_", " ")


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("source", type=pathlib.Path,
                    help="Deliveroo 'Items Sold' .xlsx from the client")
    ap.add_argument("--out-dir", type=pathlib.Path, default=None,
                    help="Output directory (default: output/blackbear/<today>)")
    ap.add_argument("--mapping", type=pathlib.Path, default=DEFAULT_MAPPING,
                    help="Deliveroo restaurant → Supy branch CSV")
    ap.add_argument("--vat-rate", type=float, default=0.20,
                    help="VAT rate used to derive sales incl. tax (default: 0.20)")
    ap.add_argument("--year", type=int, default=None,
                    help="Year for the tab dates (default: inferred from today)")
    ap.add_argument("--one-file-per-date", action="store_true",
                    help="Write one file per branch per date instead of one per branch")
    ap.add_argument("--email", metavar="ADDR", action="append", default=None,
                    help="Email the combined workbook to ADDR (repeatable). "
                         "Requires GMAIL_USER / GMAIL_APP_PASSWORD in .env")
    args = ap.parse_args()

    if not args.source.exists():
        print(f"Source file not found: {args.source}", file=sys.stderr)
        return 1
    if not 0 <= args.vat_rate < 1:
        print(f"--vat-rate must be between 0 and 1, got {args.vat_rate}", file=sys.stderr)
        return 1

    today = datetime.now().date()
    out_dir = args.out_dir or (pathlib.Path("output") / "blackbear" / today.isoformat())

    try:
        combined, rows, span = convert(args.source, out_dir, args.mapping,
                                       args.vat_rate, args.year,
                                       args.one_file_per_date, today)
        if args.email:
            branches = len([p for p in out_dir.glob("*.xlsx")
                            if not p.name.startswith("_")]) - 1
            send_email(combined, args.email, rows, span, branches,
                       out_dir / "_report.txt")
        return 0
    except ConvertError as exc:
        print(f"Error: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
